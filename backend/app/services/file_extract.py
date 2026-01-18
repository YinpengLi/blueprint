from typing import Optional
import io
from pypdf import PdfReader
from docx import Document
import openpyxl

def extract_text(filename: str, data: bytes) -> Optional[str]:
    name = (filename or "").lower()
    try:
        if name.endswith(".pdf"):
            reader = PdfReader(io.BytesIO(data))
            return "\n".join([(p.extract_text() or "") for p in reader.pages]).strip() or None
        if name.endswith(".docx"):
            doc = Document(io.BytesIO(data))
            return "\n".join([p.text for p in doc.paragraphs]).strip() or None
        if name.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            out = []
            for ws in wb.worksheets[:5]:
                out.append(f"[Sheet] {ws.title}")
                max_r = min(200, ws.max_row or 1)
                for row in ws.iter_rows(min_row=1, max_row=max_r, values_only=True):
                    out.append("\t".join("" if v is None else str(v) for v in row[:30]))
            return "\n".join(out).strip() or None
    except Exception:
        return None
    return None
